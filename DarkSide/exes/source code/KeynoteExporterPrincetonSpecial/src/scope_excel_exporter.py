"""
Scope-based Excel Exporter for KeynoteExporter PrincetonSpecial

This module handles PrincetonSpecial scope-based Excel export functionality, creating separate
Exterior and Interior Excel files based on scope tags from keynote data.
"""

import os
import logging
from typing import List, Dict, Tuple, Optional, cast
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from .keynote_data import KeynoteData
from .keynote_config import KeynoteConfig

logger = logging.getLogger(__name__)


class ScopeExcelExporter:
    TOTAL_COLUMNS = 11

    """
    Handles PrincetonSpecial scope-based Excel export for keynote data.
    Creates separate Exterior and Interior Excel files based on scope assignments.
    """
    
    def __init__(self, config: KeynoteConfig):
        """
        Initialize the scope-based Excel exporter.
        
        Args:
            config: KeynoteConfig object containing column definitions
        """
        self.config = config
        self.interior_scope_columns = config.get_interior_scope_columns()
        self.exterior_scope_columns = config.get_exterior_scope_columns()
        
    def export_scope_based_excel(self, keynote_data: List[KeynoteData], output_dir: str, original_excel_name: Optional[str] = None) -> Tuple[Optional[str], Optional[str]]:
        """
        Export keynote data to separate Exterior and Interior Excel files based on scope.
        
        Args:
            keynote_data: List of KeynoteData objects
            output_dir: Directory to save the Excel files
            original_excel_name: Name of the original Excel file (without extension)
            
        Returns:
            Tuple of (exterior_file_path, interior_file_path)
        """
        logger.info(f"Starting scope-based Excel export for {len(keynote_data)} keynote entries")
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Separate data by scope
        exterior_data = []
        interior_data = []
        
        for keynote in keynote_data:
            # Check if keynote has any exterior scope assignments
            exterior_scopes = keynote.get_exterior_scope_fields()
            if any(exterior_scopes.values()):
                exterior_data.append(keynote)
            
            # Check if keynote has any interior scope assignments
            interior_scopes = keynote.get_interior_scope_fields()
            if any(interior_scopes.values()):
                interior_data.append(keynote)
        
        logger.info(f"Found {len(exterior_data)} exterior entries and {len(interior_data)} interior entries")
        
        # Prepare file paths if needed
        exterior_path: Optional[str] = None
        interior_path: Optional[str] = None
        if original_excel_name:
            exterior_filename = f"{original_excel_name}_Exterior_Index.xlsx"
            interior_filename = f"{original_excel_name}_Interior_Index.xlsx"
        else:
            exterior_filename = "Exterior_Keynotes_Export.xlsx"
            interior_filename = "Interior_Keynotes_Export.xlsx"
        
        if exterior_data:
            exterior_path = os.path.join(output_dir, exterior_filename)
            self._create_excel_file(exterior_data, exterior_path, "Exterior")
            logger.info(f"Created exterior Excel file: {exterior_path}")
        else:
            logger.info("No exterior data found, skipping exterior file creation")
            
        if interior_data:
            interior_path = os.path.join(output_dir, interior_filename)
            self._create_excel_file(interior_data, interior_path, "Interior")
            logger.info(f"Created interior Excel file: {interior_path}")
        else:
            logger.info("No interior data found, skipping interior file creation")
        
        return exterior_path, interior_path
    
    def _create_excel_file(self, keynote_data: List[KeynoteData], file_path: str, scope_type: str) -> None:
        """
        Create an Excel file with scope-based grouping.
        
        Args:
            keynote_data: List of KeynoteData objects for this scope
            file_path: Path to save the Excel file
            scope_type: "Exterior" or "Interior"
        """
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.title = scope_type
        
        # Define styling
        header_font = Font(bold=True, size=12)
        group_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="C8C8C8", end_color="C8C8C8", fill_type="solid")
        
        current_row = 1
        
        # Create headers
        current_row = self._create_headers(ws, current_row, header_font, header_fill, thin_border)
        
        # Group data by scope tags
        scope_groups = self._group_by_scope_tags(keynote_data, scope_type)
        
        # Create content for each scope group
        for scope_tag, entries in scope_groups.items():
            if not entries:
                continue
                
            # Add spacing before each group (except the first one)
            if current_row > 3:  # Not the first group
                current_row += 1  # Add 1 empty row before each group
            
            # Add scope group header
            current_row = self._add_scope_group_header(ws, current_row, scope_tag, group_font, thin_border)
            
            last_division = None
            last_section = None

            # Add entries for this scope
            for entry in entries:
                division_key = (
                    entry.get_string_field('DIVISION #'),
                    entry.get_string_field('DIVISION NAME')
                )
                section_key = (
                    entry.get_string_field('SECTION #'),
                    entry.get_string_field('SECTION NAME')
                )

                if division_key != last_division:
                    current_row = self._add_division_header_row(ws, current_row, division_key, thin_border)
                    last_division = division_key
                    # reset section so new division always adds section header next
                    last_section = None

                if section_key != last_section:
                    current_row = self._add_section_header_row(ws, current_row, section_key, thin_border)
                    last_section = section_key

                current_row = self._add_keynote_entry(ws, current_row, entry, thin_border)
                current_row += 1  # Add spacing between entries
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
        # Save the file
        wb.save(file_path)
        logger.info(f"Saved {scope_type} Excel file with {len(keynote_data)} entries to {file_path}")
    
    def _create_headers(self, ws, start_row: int, header_font: Font, header_fill: PatternFill, border: Border) -> int:
        """
        Create a single header row for the worksheet.

        Args:
            ws: Worksheet object
            start_row: Starting row number
            header_font: Font for headers
            header_fill: Fill color for headers
            border: Border style

        Returns:
            Next row number after headers
        """
        headers = [
            "SPEC NUMBER",
            "KEYNOTE ID",
            "KEYNOTE DESCRIPTION",
            "SOURCE (OR ARCHITECT APPROVED EQUAL)",
            "PRODUCT",
            "CAT. NO.",
            "COLOR",
            "FINISH",
            "SIZE",
            "CONTACT",
            "REMARKS",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        return start_row + 1
    
    def _group_by_scope_tags(self, keynote_data: List[KeynoteData], scope_type: str) -> Dict[str, List[KeynoteData]]:
        """
        Group keynote data by scope tags.
        
        Args:
            keynote_data: List of KeynoteData objects
            scope_type: "Exterior" or "Interior"
            
        Returns:
            Dictionary mapping scope tags to lists of keynote entries
        """
        scope_columns = self.exterior_scope_columns if scope_type == "Exterior" else self.interior_scope_columns
        groups = {}
        
        # Initialize groups for each scope tag
        for scope_tag in scope_columns:
            groups[scope_tag] = []
        
        # Group entries by their scope assignments
        for keynote in keynote_data:
            if scope_type == "Exterior":
                scope_fields = keynote.get_exterior_scope_fields()
            else:
                scope_fields = keynote.get_interior_scope_fields()
            
            # Add to groups for each True scope assignment
            for scope_tag, is_active in scope_fields.items():
                if is_active:
                    groups[scope_tag].append(keynote)
        
        return groups
    
    def _add_scope_group_header(self, ws, row: int, scope_tag: str, font: Font, border: Border) -> int:
        """
        Add a scope group header row without merged cells to avoid corruption.

        Args:
            ws: Worksheet object
            row: Row number
            scope_tag: Scope tag name
            font: Font for the header
            border: Border style

        Returns:
            Next row number
        """
        # Add scope tag header without merging cells
        cell = ws.cell(row=row, column=1, value=scope_tag)
        cell.font = font
        # No border for group header cells
        cell.alignment = Alignment(horizontal='left', vertical='center')

        # Add background fill to make it stand out as a group header
        from openpyxl.styles import PatternFill
        header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        cell.fill = header_fill

        # Fill the rest of the row with the same background for visual consistency
        for col in range(2, self.TOTAL_COLUMNS + 1):  # Fill remaining columns
            fill_cell = ws.cell(row=row, column=col)
            fill_cell.fill = header_fill
            # No border for group header cells

        return row + 1
    
    def _add_keynote_entry(self, ws, row: int, keynote: KeynoteData, border: Border) -> int:
        """
        Add a keynote entry row.
        
        Args:
            ws: Worksheet object
            row: Row number
            keynote: KeynoteData object
            border: Border style
            
        Returns:
            Next row number
        """
        # Define the columns to include
        columns = [
            ("Spec Number", keynote.get_string_field('SECTION #')),
            ("KEYNOTE #", keynote.get_string_field('KEYNOTE #')),
            ("KEYNOTE DESCRIPTION", keynote.get_string_field('KEYNOTE DESCRIPTION')),
            ("SOURCE", keynote.get_string_field('SOURCE')),
            ("PRODUCT", keynote.get_string_field('PRODUCT')),
            ("CAT. NO.", keynote.get_string_field('CAT. NO.')),
            ("COLOR", keynote.get_string_field('COLOR')),
            ("FINISH", keynote.get_string_field('FINISH')),
            ("SIZE", keynote.get_string_field('SIZE')),
            ("CONTACT", keynote.get_string_field('CONTACT')),
            ("REMARKS", keynote.get_string_field('REMARKS'))
        ]
        
        for col, (header, value) in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col, value=value if col != 1 else None)
            cell.border = border
            if header == "KEYNOTE DESCRIPTION":
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            elif col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center')
        
        return row

    def _style_header_row(self, ws, row: int, border: Border, fill_color: str) -> None:
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col in range(1, self.TOTAL_COLUMNS + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.fill = header_fill
            if col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center')

    def _add_division_header_row(self, ws, row: int, division_key: tuple, border: Border) -> int:
        division_number, division_name = division_key
        if not division_number and not division_name:
            return row
        division_number = division_number or ""
        division_name = division_name or ""

        division_fill = "BFBFBF"  # Darker grey for division rows
        self._style_header_row(ws, row, border, division_fill)
        division_fill_style = PatternFill(start_color=division_fill, end_color=division_fill, fill_type="solid")
        spec_cell = ws.cell(row=row, column=1, value=division_number)
        spec_cell.alignment = Alignment(horizontal='center', vertical='center')
        spec_cell.fill = division_fill_style
        # Merge columns B and C for longer division titles
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        desc_cell = ws.cell(row=row, column=2, value=f"DIVISION {division_number} - {division_name}".strip(" -"))
        desc_cell.alignment = Alignment(horizontal='left', vertical='center')
        desc_cell.fill = division_fill_style
        return row + 1

    def _add_section_header_row(self, ws, row: int, section_key: tuple, border: Border) -> int:
        section_number, section_name = section_key
        if not section_number and not section_name:
            return row
        section_number = section_number or ""
        section_name = section_name or ""

        section_fill = "E3E3E3"  # Lighter grey for section rows
        self._style_header_row(ws, row, border, section_fill)
        section_fill_style = PatternFill(start_color=section_fill, end_color=section_fill, fill_type="solid")
        spec_cell = ws.cell(row=row, column=1, value=section_number)
        spec_cell.alignment = Alignment(horizontal='center', vertical='center')
        spec_cell.fill = section_fill_style
        # Merge columns B and C for longer section titles
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        desc_cell = ws.cell(row=row, column=2, value=f"SECTION {section_number} - {section_name}".strip(" -"))
        desc_cell.alignment = Alignment(horizontal='left', vertical='center')
        desc_cell.fill = section_fill_style
        return row + 1
    
    def _auto_adjust_columns(self, ws) -> None:
        """
        Auto-adjust column widths for better readability.
        
        Args:
            ws: Worksheet object
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width


def export_keynote_data_to_scope_excel(keynote_data: List[KeynoteData], config: KeynoteConfig, output_dir: str, original_excel_name: Optional[str] = None) -> Tuple[Optional[str], Optional[str]]:
    """
    Export keynote data to separate Exterior and Interior Excel files based on scope.
    
    Args:
        keynote_data: List of KeynoteData objects
        config: KeynoteConfig object
        output_dir: Directory to save the Excel files
        original_excel_name: Name of the original Excel file (without extension)
        
    Returns:
        Tuple of (exterior_file_path, interior_file_path)
    """
    exporter = ScopeExcelExporter(config)
    return exporter.export_scope_based_excel(keynote_data, output_dir, original_excel_name)
