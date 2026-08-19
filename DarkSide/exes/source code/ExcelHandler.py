import _Exe_Util
import openpyxl
import openpyxl.utils
import openpyxl.styles
import requests
from io import BytesIO
import os
import traceback
import datetime
import time
import xlrd  # Added for .xls support

try:
    import pythoncom  # type: ignore
    from win32com.client import DispatchEx, constants  # type: ignore
    _HAS_PYWIN32 = True
except Exception:
    _HAS_PYWIN32 = False

DEBUG = False
class ExcelHandler:
    """
    A handler for processing Excel files based on specified modes.
    """

    def __init__(self) -> None:
        """
        Initializes the ExcelHandler with job data and output storage.
        """
        if DEBUG:
            self.job_data = _Exe_Util.get_data("DEBUGER_excel_handler_input")
        else:
            self.job_data = _Exe_Util.get_data("excel_handler_input")
        self.out = {}

    @_Exe_Util.try_catch_error
    def run(self):
        """
        Executes the appropriate action based on the mode specified in job data.

        Any unhandled exception inside the dispatched reader/writer is captured
        and written back to ``excel_handler_input`` as ``status="error"`` with
        the traceback in the ``error`` field, so the IronPython-side caller can
        surface the real cause to the user instead of waiting for a timeout.
        """
        mode = self.job_data.get("mode", None)
        try:
            if mode == "read":
                ExcelReader().run()
            elif mode == "write":
                ExcelWriter().run()
            elif mode == "update":
                ExcelUpdater().run()
            else:
                self._signal_error("Unknown mode {!r} or no job data.".format(mode))
        except Exception as exc:
            err = "{}\n{}".format(exc, traceback.format_exc())
            self._signal_error(err)

    def _signal_error(self, message):
        """Surface an error to the caller via the input-channel status field."""
        print("ExcelHandler error: {}".format(message))
        try:
            _Exe_Util.set_data({}, "excel_handler_output")
            _Exe_Util.set_data(
                {"status": "error", "error": str(message)},
                "excel_handler_input",
            )
        except Exception as inner:
            print("ExcelHandler: failed to write error status: {}".format(inner))

    def dump_data(self):
        """
        Dumps the processed data to a specified output file.

        Any non-fatal warnings collected during the read (e.g. Conditional
        Formatting detected on the sheet, theme color resolution failed)
        are attached to the status payload so the caller can surface them.
        """
        _Exe_Util.set_data(self.out, "excel_handler_output")
        status_payload = {"status": "done"}
        warnings = getattr(self, "warnings", None)
        if warnings:
            status_payload["warnings"] = list(warnings)
        _Exe_Util.set_data(status_payload, "excel_handler_input")

        if DEBUG:
            file_path = self.job_data.get("filepath", None)
            if file_path:
                os.startfile(file_path)

class ExcelReader(ExcelHandler):
    """
    A class to read Excel files and extract data, preserving cell values and colors.
    Supports both .xlsx and .xls formats from local files and HTTP/HTTPS URLs.
    
    Both formats use identical return structures with 1-based indexing for rows and columns,
    ensuring consistent behavior regardless of the source file format.
    
    In the returned dictionary, Excel's top-left cell (A1) will have the key "1,1",
    representing row 1, column 1. All cells follow this pattern of "row,column" with
    1-based indexing that matches Excel's native cell addressing.
    """

    def __init__(self) -> None:
        """
        Initializes the ExcelReader by calling the parent constructor.
        """
        super().__init__()
        # Non-fatal hints accumulated during the read (Conditional Formatting,
        # theme color resolution failures, etc). Surfaced via dump_data so the
        # caller can show them next to "No valid color entries" failures.
        self.warnings = []
        self._theme_warn_count = 0

    def run(self):
        """
        Reads an Excel file from a specified path or URL, handling both .xlsx and .xls formats.
        Supports reading from local files and HTTP/HTTPS URLs.
        """
        file_path = self.job_data.get("filepath", None)
        if file_path:
            # Handle HTTP URLs
            if file_path.startswith("http"):
                response = requests.get(file_path)
                file_content = BytesIO(response.content)
                
                # Determine file type based on URL extension
                if file_path.lower().endswith(".xlsx"):
                    self.out = self.read_xlsx(file_content)
                elif file_path.lower().endswith(".xls"):
                    self.out = self.read_xls(file_content)
                else:
                    # Try xlsx format by default
                    try:
                        self.out = self.read_xlsx(file_content)
                    except Exception:
                        # Fall back to xls format
                        file_content.seek(0)  # Reset the BytesIO position
                        self.out = self.read_xls(file_content)
            
            # Handle local files. Will not skip any cell even if it is empty. Becasue fore color loading i need to have all cell even empty cell, which will have cllor info
            else:
                if file_path.lower().endswith(".xlsx"):
                    self.out = self.read_xlsx(file_path)
                elif file_path.lower().endswith(".xls"):
                    self.out = self.read_xls(file_path)
                else:
                    print("ExcelReader: Unsupported file format. Please use .xlsx or .xls files.")

        _Exe_Util.set_data(self.out, "DEBUGER_excel_handler_output")

        self.dump_data()

    def read_xlsx(self, file_content):
        """
        Reads data from an Excel .xlsx file, extracting cell values and colors.
        
        Args:
            file_content (str or BytesIO): Either a file path to a local Excel file
                                         or a BytesIO object containing file data from HTTP.

        Returns:
            dict: A dictionary with cell positions as keys (format: "row,column") and 
                 their values and colors. Both row and column are 1-based indices.
                 For example, Excel cell A1 is represented as key "1,1".
                 Empty cells with formatting are also included.
        """
        # Note: Not using read_only=True to allow proper theme color resolution
        # data_only=True ensures we get calculated values instead of formulas
        workbook = openpyxl.load_workbook(file_content, data_only=True, read_only=False, rich_text=True)
        requested_sheet = self.job_data.get("worksheet")
        sheet_name = requested_sheet if requested_sheet else workbook.sheetnames[0]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                "Worksheet '{}' not found in workbook. Available worksheets: {}".format(
                    sheet_name, list(workbook.sheetnames)
                )
            )
        sheet = workbook[sheet_name]

        # Surface a hint if the worksheet uses Conditional Formatting -- those
        # visually-painted cells look correct in Excel but openpyxl can only
        # see the underlying (often-empty) fill, which is the #1 cause of
        # "No valid color entries" surprises.
        try:
            cf_rule_count = sum(len(rules) for rules in sheet.conditional_formatting._cf_rules.values())
        except Exception:
            cf_rule_count = 0
        if cf_rule_count:
            self.warnings.append(
                "Worksheet '{}' has {} Conditional Formatting rule(s). "
                "openpyxl cannot read CF-painted colors. If your color column "
                "is colored via CF, replace it with a direct Fill Color.".format(
                    sheet_name, cf_rule_count
                )
            )

        data = {}
        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                # Handle both regular cells and EmptyCell objects
                try:
                    row_num = cell.row
                    col_num = cell.column
                except AttributeError:
                    # EmptyCell doesn't have row/column attributes, use the iteration indices
                    row_num = row_idx
                    col_num = col_idx
                
                key = "{},{}".format(row_num, col_num)  # Using format instead of f-string per instructions
                value = self._format_cell_value(cell.value)
                data[key] = {
                    'color': self.get_xlsx_cell_color(cell),
                    'value': value
                }
        return data

    def read_xls(self, file_content):
        """
        Reads data from an Excel .xls file, extracting cell values and colors.
        
        Args:
            file_content (str or BytesIO): Either a file path to a local Excel file
                                         or a BytesIO object containing file data from HTTP.

        Returns:
            dict: A dictionary with cell positions as keys (format: "row,column") and 
                 their values and colors. Both row and column are 1-based indices.
                 For example, Excel cell A1 is represented as key "1,1".
        """
        workbook = xlrd.open_workbook(file_content, formatting_info=True)
        sheet_name = self.job_data.get("worksheet", None)

        # Get the sheet either by specified name or the first sheet.
        # If a name is given but does not match, surface the mismatch instead
        # of silently falling back to sheet 0 (which previously caused the
        # caller to read the wrong tab and report "No valid color entries").
        if sheet_name and sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
        elif sheet_name:
            raise ValueError(
                "Worksheet '{}' not found in .xls workbook. Available worksheets: {}".format(
                    sheet_name, workbook.sheet_names()
                )
            )
        else:
            sheet = workbook.sheet_by_index(0)
            
        data = {}
        
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(row_idx, col_idx)
                
                
                # Convert datetime if needed
                cell_type = sheet.cell_type(row_idx, col_idx)
                if cell_type == xlrd.XL_CELL_DATE:
                    cell_value = xlrd.xldate.xldate_as_datetime(
                        cell_value, workbook.datemode
                    ).isoformat()
                
                # Format the value consistently
                cell_value = self._format_cell_value(cell_value)
                
                # Get cell position as key - convert to 1-based for consistency with xlsx
                # Use f-string for key format
                key = f"{row_idx + 1},{col_idx + 1}"
                
                # Get cell color
                color = self.get_xls_cell_color(workbook, sheet, row_idx, col_idx)
                
                data[key] = {
                    'color': color,
                    'value': cell_value
                }
                
        return data

    def get_xls_cell_color(self, workbook, sheet, row_idx, col_idx):
        """
        Retrieves the RGB color of a cell's fill in an .xls file.

        Args:
            workbook (xlrd.Book): The workbook containing the cell
            sheet (xlrd.Sheet): The worksheet containing the cell
            row_idx (int): Row index of the cell
            col_idx (int): Column index of the cell

        Returns:
            tuple: A tuple representing the RGB color (r, g, b), or (None, None, None) if default/no fill.
                  Values range from 0 to 255 for each component.
        """
        try:
            xf_index = sheet.cell_xf_index(row_idx, col_idx)
            xf_record = workbook.xf_list[xf_index]
            bg_pattern_index = xf_record.background.pattern_colour_index
            
            if bg_pattern_index > 0:
                # Get color from workbook color map
                color = workbook.colour_map.get(bg_pattern_index)
                if color:
                    # Convert color components from range 0-255
                    r, g, b = color[0], color[1], color[2]
                    return (r, g, b)
        except Exception:
            # If any error occurs during color extraction, return None
            pass
            
        return (None, None, None)

    def get_xlsx_cell_color(self, cell):
        """
        Retrieves the RGB color of a cell's fill in an .xlsx file.
        Supports direct RGB colors, ARGB colors, and theme colors with tint.

        Args:
            cell (openpyxl.cell): The cell from which to retrieve the color.
                                  Can be a regular cell or an EmptyCell object.

        Returns:
            tuple: A tuple representing the RGB color (r, g, b), or (None, None, None) if default/no fill.
                  Values range from 0 to 255 for each component.
        """
        try:
            if cell.fill:
                # Try to get color from fgColor first (foreground color)
                if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                    fg_color = cell.fill.fgColor
                    
                    # Check if fgColor has direct RGB
                    if hasattr(fg_color, 'rgb') and isinstance(fg_color.rgb, str):
                        color_hex = fg_color.rgb
                        if len(color_hex) == 8:  # ARGB format
                            color_hex = color_hex[2:]
                        if len(color_hex) == 6:
                            r = int(color_hex[0:2], 16)
                            g = int(color_hex[2:4], 16)
                            b = int(color_hex[4:6], 16)
                            return (r, g, b)
                    
                    # Check if fgColor is a theme color
                    if hasattr(fg_color, 'type') and fg_color.type == 'theme':
                        return self._get_theme_color_rgb(cell, fg_color)
                
                # Fall back to start_color
                if hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                    start_color = cell.fill.start_color
                    
                    # Check if this is a theme color
                    if hasattr(start_color, 'type') and start_color.type == 'theme':
                        return self._get_theme_color_rgb(cell, start_color)
                    
                    # Handle direct RGB colors
                    if hasattr(start_color, 'rgb'):
                        color = start_color.rgb
                        
                        # Handle RGB object vs string representation
                        if hasattr(color, 'rgb'):  # Newer openpyxl versions might have RGB objects
                            color = color.rgb
                            
                        # Convert string color representations (may be ARGB or RGB format)
                        if isinstance(color, str):
                            if len(color) == 8:  # ARGB format
                                color = color[2:]  # Remove alpha channel
                            if len(color) == 6:  # RGB format
                                # Parse RGB values
                                r = int(color[0:2], 16)
                                g = int(color[2:4], 16)
                                b = int(color[4:6], 16)
                                return (r, g, b)
                        
                        # In case the color is in another format but still valid
                        if hasattr(color, 'red') and hasattr(color, 'green') and hasattr(color, 'blue'):
                            try:
                                # Use getattr to avoid linter issues
                                red = getattr(color, 'red', None)
                                green = getattr(color, 'green', None)
                                blue = getattr(color, 'blue', None)
                                if red is not None and green is not None and blue is not None:
                                    return (red, green, blue)
                            except AttributeError:
                                pass
        except AttributeError:
            # EmptyCell objects may not have fill attributes
            pass
                    
        return (None, None, None)
    
    def _get_theme_color_rgb(self, cell, start_color):
        """
        Converts a theme color to RGB values.
        
        Theme colors reference the workbook's color scheme and can have a tint value
        that lightens or darkens the base theme color.
        
        Args:
            cell: The cell object (to get workbook reference)
            start_color: The Color object with theme and tint information
            
        Returns:
            tuple: RGB color as (r, g, b) or (None, None, None) if conversion fails
        """
        try:
            # Get theme index and tint
            theme_index = start_color.theme
            tint = start_color.tint if hasattr(start_color, 'tint') else 0.0
            
            # Get the workbook's theme colors
            workbook = cell.parent.parent
            
            # Try to get actual theme colors from the workbook
            base_color_hex = None
            
            # Attempt to read theme colors from workbook's theme
            if hasattr(workbook, 'loaded_theme') and workbook.loaded_theme:
                try:
                    # Parse theme XML if it's in bytes format
                    theme = workbook.loaded_theme
                    
                    # If theme is bytes (raw XML), try to parse it
                    if isinstance(theme, bytes):
                        from xml.etree import ElementTree as ET
                        # Parse the theme XML
                        root = ET.fromstring(theme)
                        
                        # Define XML namespaces
                        namespaces = {
                            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
                        }
                        
                        # Map theme index to color scheme element names
                        theme_color_map = {
                            0: 'a:lt1',  # Light 1 (Background 1)
                            1: 'a:dk1',  # Dark 1 (Text 1)
                            2: 'a:lt2',  # Light 2 (Background 2)
                            3: 'a:dk2',  # Dark 2 (Text 2)
                            4: 'a:accent1',
                            5: 'a:accent2',
                            6: 'a:accent3',
                            7: 'a:accent4',
                            8: 'a:accent5',
                            9: 'a:accent6',
                        }
                        
                        color_xpath = theme_color_map.get(theme_index)
                        if color_xpath:
                            # Find the color scheme element
                            color_scheme_path = './/a:themeElements/a:clrScheme/' + color_xpath.split(':')[1]
                            color_element = root.find(color_scheme_path, namespaces)
                            
                            if color_element is not None:
                                # Look for srgbClr (standard RGB) or sysClr (system color)
                                srgb_elem = color_element.find('.//a:srgbClr', namespaces)
                                if srgb_elem is not None and 'val' in srgb_elem.attrib:
                                    base_color_hex = srgb_elem.attrib['val']
                                else:
                                    sysclr_elem = color_element.find('.//a:sysClr', namespaces)
                                    if sysclr_elem is not None and 'lastClr' in sysclr_elem.attrib:
                                        base_color_hex = sysclr_elem.attrib['lastClr']
                    
                    # If theme has already been parsed (object with attributes)
                    elif hasattr(theme, 'themeElements') and hasattr(theme.themeElements, 'clrScheme'):
                        color_scheme = theme.themeElements.clrScheme
                        
                        # Map theme index to color scheme element
                        theme_color_map = {
                            0: 'lt1',  # Light 1 (Background 1)
                            1: 'dk1',  # Dark 1 (Text 1)
                            2: 'lt2',  # Light 2 (Background 2)
                            3: 'dk2',  # Dark 2 (Text 2)
                            4: 'accent1',
                            5: 'accent2',
                            6: 'accent3',
                            7: 'accent4',
                            8: 'accent5',
                            9: 'accent6',
                        }
                        
                        color_name = theme_color_map.get(theme_index)
                        if color_name and hasattr(color_scheme, color_name):
                            color_element = getattr(color_scheme, color_name)
                            # Try to extract RGB from the color element
                            if hasattr(color_element, 'srgbClr') and color_element.srgbClr:
                                base_color_hex = color_element.srgbClr.val
                            elif hasattr(color_element, 'sysClr') and color_element.sysClr:
                                base_color_hex = color_element.sysClr.lastClr
                except Exception as e:
                    # If reading from theme fails, will fall back to defaults
                    pass
            
            # Fall back to default Excel theme colors (Office theme) if not found
            if not base_color_hex:
                default_theme_colors = [
                    "FFFFFF",  # 0: Background 1 (White)
                    "000000",  # 1: Text 1 (Black)
                    "E7E6E6",  # 2: Background 2 (Light Gray)
                    "44546A",  # 3: Text 2 (Dark Blue Gray)
                    "4472C4",  # 4: Accent 1 (Blue)
                    "ED7D31",  # 5: Accent 2 (Orange)
                    "A5A5A5",  # 6: Accent 3 (Gray)
                    "FFC000",  # 7: Accent 4 (Gold)
                    "5B9BD5",  # 8: Accent 5 (Light Blue)
                    "70AD47",  # 9: Accent 6 (Green)
                ]
                
                if theme_index is not None and 0 <= theme_index < len(default_theme_colors):
                    base_color_hex = default_theme_colors[theme_index]
                else:
                    self._record_theme_failure(cell, "theme index {} out of range".format(theme_index))
                    return (None, None, None)

            # Clean up hex string (remove any alpha channel)
            if isinstance(base_color_hex, str):
                if len(base_color_hex) == 8:  # ARGB format
                    base_color_hex = base_color_hex[2:]
                elif len(base_color_hex) != 6:
                    self._record_theme_failure(cell, "theme hex has unexpected length: {!r}".format(base_color_hex))
                    return (None, None, None)
            else:
                self._record_theme_failure(cell, "theme color did not resolve to a hex string")
                return (None, None, None)
            
            # Parse base color
            r = int(base_color_hex[0:2], 16)
            g = int(base_color_hex[2:4], 16)
            b = int(base_color_hex[4:6], 16)
            
            # Apply tint if present
            if tint != 0:
                r, g, b = self._apply_tint_to_rgb(r, g, b, tint)
            
            return (r, g, b)
            
        except Exception as e:
            # If theme color conversion fails, return None
            self._record_theme_failure(cell, "theme parse exception: {}".format(e))
            return (None, None, None)

    def _record_theme_failure(self, cell, reason):
        """Track theme-color resolution failures so the caller can surface them.

        Only the first 5 distinct failures are reported verbatim to avoid
        flooding the caller's UI when an entire column is theme-painted.
        """
        self._theme_warn_count += 1
        if self._theme_warn_count > 5:
            return
        try:
            coord = cell.coordinate
        except Exception:
            coord = "(unknown cell)"
        self.warnings.append(
            "Theme color could not be resolved for cell {}: {}".format(coord, reason)
        )
        if self._theme_warn_count == 5:
            self.warnings.append(
                "Additional theme-color failures suppressed -- replace theme "
                "colors with direct Fill Colors in the affected column."
            )

    def _apply_tint_to_rgb(self, r, g, b, tint):
        """
        Applies a tint value to RGB colors following Excel's tint algorithm.
        
        Tint values range from -1.0 (darkest) to 1.0 (lightest):
        - Negative tint values darken the color
        - Positive tint values lighten the color
        - 0.0 means no change
        
        Args:
            r, g, b: RGB values (0-255)
            tint: Tint value (-1.0 to 1.0)
            
        Returns:
            tuple: Tinted RGB values as (r, g, b)
        """
        # Convert RGB to 0-1 range
        r_norm = r / 255.0
        g_norm = g / 255.0
        b_norm = b / 255.0
        
        # Apply tint using Excel's algorithm
        if tint < 0:
            # Darken: RGB = RGB * (1 + tint)
            r_norm = r_norm * (1 + tint)
            g_norm = g_norm * (1 + tint)
            b_norm = b_norm * (1 + tint)
        else:
            # Lighten: RGB = RGB * (1 - tint) + tint
            r_norm = r_norm * (1 - tint) + tint
            g_norm = g_norm * (1 - tint) + tint
            b_norm = b_norm * (1 - tint) + tint
        
        # Convert back to 0-255 range and clamp
        r_tinted = max(0, min(255, int(r_norm * 255)))
        g_tinted = max(0, min(255, int(g_norm * 255)))
        b_tinted = max(0, min(255, int(b_norm * 255)))
        
        return (r_tinted, g_tinted, b_tinted)

    def _format_cell_value(self, value):
        """
        Formats a cell value consistently based on its type.

        Args:
            value (any): The value to be formatted.

        Returns:
            str: The formatted value.
        """
        if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
            return value.isoformat()
        elif not isinstance(value, (str, int, float, bool)):
            return str(value)
        else:
            return value

class ExcelWriter(ExcelHandler):
    """
    A class to write data to Excel files with support for formatting.
    
    The class accepts data as a list of objects with the following attributes:
        - item: Cell content (any type)
        - row (int): Row index
        - column (int): Column index
        - is_bold (bool): If True, the text will be bold
        - is_read_only (bool): If True, the cell will be read only
        - cell_color (tuple): RGB color tuple for cell background
        - text_color (tuple): RGB color tuple for text
        - text_alignment (str): Text alignment
        - border_style (int): Border style specification
        - border_color (tuple): RGB color tuple for border
        - top_border_style (int): Top border style specification
        - side_border_style (int): Left/right border style specification
        - bottom_border_style (int): Bottom border style specification
        - col_width (float): Column width value (overrides auto-sizing)
        - font_size (int): Font size in points
        - font_name (str): Font family name
        - merge_with (list): List of (row, col) tuples to merge with this cell
        - text_wrap (bool): If True, the text will be wrapped
        - tooltip (str): Tooltip text to display when hovering over the cell
    """

    def __init__(self) -> None:
        """
        Initializes the ExcelWriter by calling the parent constructor.
        Also initializes collections to track protected and unprotected cells.
        """
        super().__init__()
        
        # Collections to track cell protection status
        self.protected_cells = []    # Cells that require password to edit
        self.unprotected_cells = []  # Cells that can be edited without password
        self.need_protection = False

    def run(self):
        """
        Creates and saves an Excel file based on the job data.
        
        This method:
        1. Processes data from job_data
        2. Creates a workbook with the specified data
        3. Applies cell formatting for each cell
        4. Applies protection settings if needed
        5. Saves the workbook to the specified file path
        """
        file_path = self.job_data.get("filepath", None)
        worksheet_name = self.job_data.get("worksheet", None)
        data_entries = self.job_data.get("data", {}).values()
        freeze_row = self.job_data.get("freeze_row", None)
        freeze_column = self.job_data.get("freeze_column", None)
        
        if not file_path or not data_entries:
            print("ExcelWriter: Missing filepath or data in job_data.")
            return
        
        try:
            # Create workbook and worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            
            # Set worksheet name
            if worksheet_name:
                worksheet.title = worksheet_name
            
            # Track column widths - used for auto-sizing columns that don't have explicit width
            column_widths = {}
            
            # Track columns with custom widths to avoid auto-adjusting them
            custom_width_columns = set()
            
            # Track cells to merge - process after all cells are created
            cells_to_merge = []

            # Process data entries
            for data_entry in data_entries:
                # Get row and column
                row = data_entry.get("row") + 1  # Convert to 1-based for openpyxl
                
                # Handle column whether it's an index or letter
                if isinstance(data_entry.get("column"), str):
                    col = openpyxl.utils.column_index_from_string(data_entry.get("column"))
                else:
                    col = data_entry.get("column") + 1  # Convert to 1-based for openpyxl
                
                # Get cell value
                value = data_entry.get("item")
                
                # Check if this cell has a custom width setting
                if data_entry.get("col_width"):
                    custom_width_columns.add(col)
                
                # Write the value to the cell
                cell = worksheet.cell(row=row, column=col, value=value)
                
                # Apply cell formatting
                self._apply_cell_formatting(cell, data_entry)
                
                # Check if this cell needs to be merged with others
                if data_entry.get("merge_with"):
                    start_cell = cell.coordinate
                    # Calculate the end coordinate by finding the max row and column
                    merge_cells = data_entry.get("merge_with")
                    if merge_cells and isinstance(merge_cells, list):
                        print ("ExcelWriter: Merging cell desired, {} with {}".format(start_cell, merge_cells))
                        # Convert all merge cells to coordinates
                        merge_coordinates = []
                        for merge_cell in merge_cells:
                            if isinstance(merge_cell, list) and len(merge_cell) == 2:
                                merge_row, merge_col_letter = merge_cell
                                # Handle row
                                if isinstance(merge_row, int):
                                    merge_row = merge_row + 1  # Convert to 1-based
                                
                                # Handle column
                                if isinstance(merge_col_letter, str):
                                    merge_col = openpyxl.utils.column_index_from_string(merge_col_letter)
                                else:
                                    merge_col = merge_col_letter + 1  # Convert to 1-based
                                
                                merge_coordinates.append((merge_row, merge_col))
                                print ("ExcelWriter: Merging cell task added {} with {}".format(start_cell, (merge_row, merge_col)))
                        
                        if merge_coordinates:
                            # Calculate the merge range
                            max_row = max([cell_coord[0] for cell_coord in merge_coordinates])
                            max_col = max([cell_coord[1] for cell_coord in merge_coordinates])
                            end_cell = worksheet.cell(row=max_row, column=max_col).coordinate
                            
                            # Store merge information for later processing
                            cells_to_merge.append((start_cell, end_cell))
                            
                            # Center the text in the merged cell
                            if not cell.alignment:
                                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                            else:
                                # Preserve other alignment settings while setting horizontal and vertical centering
                                alignment_props = {
                                    'horizontal': 'center',
                                    'vertical': 'center',
                                    'textRotation': cell.alignment.textRotation,
                                    'wrapText': cell.alignment.wrapText,
                                    'shrinkToFit': cell.alignment.shrinkToFit,
                                    'indent': cell.alignment.indent,
                                    'relativeIndent': cell.alignment.relativeIndent,
                                    'justifyLastLine': cell.alignment.justifyLastLine,
                                    'readingOrder': cell.alignment.readingOrder
                                }
                                cell.alignment = openpyxl.styles.Alignment(**alignment_props)
                
                # Track maximum content length for auto-width calculation
                # (only for columns without custom width)
                if col not in custom_width_columns:
                    if col not in column_widths:
                        column_widths[col] = 0
                    # Calculate width based on content length
                    content_length = len(str(value)) if value is not None else 0
                    column_widths[col] = max(column_widths[col], content_length)

            # Apply cell merges after all cells are created
            for start_cell, end_cell in cells_to_merge:
                merge_range = "{}:{}".format(start_cell, end_cell)
                worksheet.merge_cells(merge_range)
                print("ExcelWriter: Merged cell range {}".format(merge_range))

            # set all cell level protection as unlocked
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.coordinate in self.protected_cells:
                        continue
                    cell.protection = openpyxl.styles.Protection(locked=False)
            
            # Adjust column widths only for columns that don't have custom widths
            for col, width in column_widths.items():
                if col not in custom_width_columns:
                    # Apply a width multiplier for better readability
                    adjusted_width = width * 1.1
                    col_letter = openpyxl.utils.get_column_letter(col)
                    worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Apply freeze panes if specified
            self._apply_freeze_panes(worksheet, freeze_row, freeze_column)

            # Apply protection settings if needed
            if self.need_protection:
                self._apply_protection_settings(worksheet)

            # Save the workbook
            workbook.save(file_path)
            workbook.close()
            print("ExcelWriter: File successfully saved to '{}'".format(file_path))

            finalize_method = self._auto_finalize_excel(file_path)
            if finalize_method:
                print("ExcelWriter: Workbook finalized via {} automation".format(finalize_method))
            else:
                print("ExcelWriter: Skipped Excel automation finalize ({}).".format(file_path))
            
            # Set the output status
            self.out = {"status": "success", "filepath": file_path}
            
        except Exception as e:
            error_message = "{}".format(e)
            error_traceback = traceback.format_exc()
            print("ExcelWriter: Error writing Excel file - {}".format(error_message))
            _Exe_Util.messenger("Cannot save Excel [ {} ]\nReason: {}\nTraceback: {}".format(file_path, error_message, error_traceback))
            self.out = {"status": "error", "message": error_message, "traceback": error_traceback}
        
        self.dump_data()
    
    def _auto_finalize_excel(self, file_path):
        """Attempt to re-save the workbook through Excel automation to ensure theme styles persist."""
        if not _HAS_PYWIN32:
            print("ExcelWriter: pywin32 not available; skipping finalize for {}".format(file_path))
            return None

        try:
            pythoncom.CoInitialize()
            excel = DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            try:
                time.sleep(0.25)
                workbook = excel.Workbooks.Open(
                    file_path,
                    UpdateLinks=constants.xlUpdateLinksNever,
                    ReadOnly=False,
                    IgnoreReadOnlyRecommended=True
                )
                workbook.Save()
                workbook.Close(SaveChanges=True)
                return "pywin32"
            finally:
                excel.Quit()
                pythoncom.CoUninitialize()
        except Exception as exc:
            print("ExcelWriter: pywin32 finalize failed for {} - {}".format(file_path, exc))
        return None

    def _apply_protection_settings(self, worksheet):
        return
        """
        Applies protection settings to the worksheet using standard cell protection.
        
        This method applies protection in a way that:
        - Locked cells require a password to edit
        - Unlocked cells can be edited without a password
        - Only applies protection to specific cells, not the entire sheet
        
        Args:
            worksheet (openpyxl.Worksheet): The worksheet to apply protection settings.
        """
        print("ExcelWriter: Applying protection settings...")
        
        # Step 1: Make sure all protection is correctly set at the cell level
        # (This should already be done in _apply_cell_formatting)
        
        # Step 2: Apply sheet protection with appropriate settings
        # This enables the locked/unlocked cell attributes to take effect
        sheet_protection = openpyxl.worksheet.protection.SheetProtection(
            sheet=True,                # Enable protection
            password=None,         # Set password for protected areas
            formatCells=False,         # Disallow formatting cells
            formatColumns=False,       # Disallow formatting columns
            formatRows=False,          # Disallow formatting rows
            insertColumns=False,       # Disallow inserting columns
            insertRows=False,          # Disallow inserting rows
            insertHyperlinks=False,    # Disallow inserting hyperlinks
            deleteColumns=False,       # Disallow deleting columns
            deleteRows=False,          # Disallow deleting rows
            sort=False,                # Disallow sorting
            autoFilter=False,          # Disallow filters
            pivotTables=False,         # Disallow pivot tables
            selectLockedCells=True,    # Allow selecting locked cells
            selectUnlockedCells=True   # Allow selecting unlocked cells
        )
        
        # Apply the protection to the worksheet
        worksheet.protection = sheet_protection
        
        print("ExcelWriter: Protected {} locked cells while allowing {} unlocked cells to be edited freely".format(
            len(self.protected_cells), len(self.unprotected_cells)))
        print("ExcelWriter: Protection settings applied successfully")

    def _apply_cell_formatting(self, cell, data):
        """
        Applies formatting to a cell based on the data object attributes.
        
        Supported formatting attributes:
        - cell_color: RGB tuple for cell background
        - text_color: RGB tuple for text color
        - is_bold: Boolean for bold text
        - border_style: Integer index for border style
        - border_color: RGB tuple for border color
        - top_border_style: Integer index for top border style
        - bottom_border_style: Integer index for bottom border style
        - side_border_style: Integer index for side borders style
        - is_read_only: Boolean for cell protection
        - col_width: Float for column width
        - font_size: Integer for font size
        - font_name: String for font name
        - text_alignment: String for text alignment
        - text_wrap: Boolean for text wrap
        - tooltip: String for tooltip text when hovering over the cell
        
        Args:
            cell (openpyxl.Cell): The cell to format
            data (object): Object with formatting attributes
        """
        # Handle cell background color
        if data.get("cell_color"):
            r, g, b = data.get("cell_color")
            if all(x is not None for x in [r, g, b]):
                rgb_str = "{:02x}{:02x}{:02x}".format(r, g, b)
                fill = openpyxl.styles.PatternFill(
                    start_color=rgb_str,
                    end_color=rgb_str,
                    fill_type="solid"
                )
                cell.fill = fill
        
        # Create a dict of font properties
        font_props = {}

        # Add text color if present
        if data.get("text_color"):
            r, g, b = data.get("text_color")
            if all(x is not None for x in [r, g, b]):
                rgb_str = "{:02x}{:02x}{:02x}".format(r, g, b)
                font_props["color"] = rgb_str

        # Add bold if present
        if data.get("is_bold"):
            font_props["bold"] = data.get("is_bold")
        
        # Add font size if present
        if data.get("font_size"):
            font_size = data.get("font_size")
            if isinstance(font_size, (int, float)):
                font_props["size"] = font_size
            
        # Add font name if present
        if data.get("font_name"):
            font_name = data.get("font_name")
            if isinstance(font_name, str) and font_name.strip():
                font_props["name"] = font_name

        # Correctly add text alignment if present
        if data.get("text_alignment"):
            text_alignment = data.get("text_alignment")
            if isinstance(text_alignment, str):
                # Create or update alignment object
                alignment_props = {}
                if cell.alignment:
                    # Preserve existing alignment properties
                    alignment_props = {
                        'vertical': cell.alignment.vertical,
                        'textRotation': cell.alignment.textRotation,
                        'wrapText': cell.alignment.wrapText,
                        'shrinkToFit': cell.alignment.shrinkToFit,
                        'indent': cell.alignment.indent,
                        'relativeIndent': cell.alignment.relativeIndent,
                        'justifyLastLine': cell.alignment.justifyLastLine,
                        'readingOrder': cell.alignment.readingOrder
                    }
                # Set the horizontal alignment property
                alignment_props['horizontal'] = text_alignment
                cell.alignment = openpyxl.styles.Alignment(**alignment_props)
        
        # Apply combined font properties in one go
        if font_props:
            # Preserve existing font properties that we're not changing
            current_font = cell.font
            if current_font:
                # Copy existing font properties that aren't being explicitly set
                for attr in ['italic', 'underline', 'strike', 'vertAlign', 'charset', 'scheme']:
                    if attr not in font_props and hasattr(current_font, attr):
                        font_props[attr] = getattr(current_font, attr)
            
            cell.font = openpyxl.styles.Font(**font_props)
        
        # Handle column width and text wrapping
        if data.get("col_width"):
            col_width = data.get("col_width")
            if isinstance(col_width, (int, float)) and col_width > 0:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                worksheet = cell.parent
                
                # Set column width
                worksheet.column_dimensions[col_letter].width = col_width
                
                # If text is longer than column width, enable text wrapping
                if cell.value and isinstance(cell.value, str):
                    estimated_text_width = len(cell.value) * 1.2  # rough estimate
                    if estimated_text_width > col_width:
                        if not cell.alignment:
                            cell.alignment = openpyxl.styles.Alignment()
                        cell.alignment = openpyxl.styles.Alignment(
                            wrap_text=True, 
                            vertical='top',
                            horizontal=cell.alignment.horizontal if cell.alignment else 'general'
                        )
        
        # Handle read-only property separately through cell protection
        if data.get("is_read_only"):
            cell.protection = openpyxl.styles.Protection(locked=True)
            if cell.coordinate not in self.protected_cells:
                self.protected_cells.append(cell.coordinate)  # Add to protected cells
            self.need_protection = True
        else:
            cell.protection = openpyxl.styles.Protection(locked=False)
            if cell.coordinate not in self.unprotected_cells:
                self.unprotected_cells.append(cell.coordinate)  # Add to unprotected cells
        
        # Handle text wrapping
        if data.get("text_wrap"):
            if not cell.alignment:
                cell.alignment = openpyxl.styles.Alignment()
            cell.alignment = openpyxl.styles.Alignment(
                wrap_text=True,
                vertical=cell.alignment.vertical if cell.alignment else 'top',
                horizontal=cell.alignment.horizontal if cell.alignment else 'general'
            )
        
        # Handle borders
        border_sides = {}
        border_color = None
        
        # Get border color if specified
        if data.get("border_color"):
            r, g, b = data.get("border_color")
            if all(x is not None for x in [r, g, b]):
                border_color = "{:02x}{:02x}{:02x}".format(r, g, b)
        
        # Regular borders
        if data.get("border_style"):
            border_style = self._convert_border_style(data.get("border_style"))
            
            # Set all sides with the same style
            for side in ['left', 'right', 'top', 'bottom']:
                border_sides[side] = openpyxl.styles.Side(
                    border_style=border_style,
                    color=border_color
                )
        
        # Top border (overrides regular border if set)
        if data.get("top_border_style"):
            border_style = self._convert_border_style(data.get("top_border_style"))
            border_sides['top'] = openpyxl.styles.Side(
                border_style=border_style,
                color=border_color
            )
        
        # Bottom border (overrides regular border if set)
        if data.get("bottom_border_style"):
            border_style = self._convert_border_style(data.get("bottom_border_style"))
            border_sides['bottom'] = openpyxl.styles.Side(
                border_style=border_style,
                color=border_color
            )
        # Side borders (overrides regular border if set)
        if data.get("side_border_style"):
            border_style = self._convert_border_style(data.get("side_border_style"))
            border_sides['left'] = openpyxl.styles.Side(
                border_style=border_style,
                color=border_color
            )
            border_sides['right'] = openpyxl.styles.Side(
                border_style=border_style,
                color=border_color
            )
        
        # Apply borders if any were defined
        if border_sides:
            cell.border = openpyxl.styles.Border(**border_sides)
        
        # Handle tooltip functionality
        if data.get("tooltip"):
            tooltip_data = data.get("tooltip")
            
            # Handle both string and dict tooltip formats
            if isinstance(tooltip_data, dict):
                # Dict format: {'title': 'Title', 'content': 'Content'}
                tooltip_title = tooltip_data.get("title", "Tooltip")
                tooltip_content = tooltip_data.get("content", "")
            elif isinstance(tooltip_data, str) and tooltip_data.strip():
                # String format: use as content with default title
                tooltip_title = "Tooltip"
                tooltip_content = tooltip_data
            else:
                tooltip_title = None
                tooltip_content = None
            
            if tooltip_content and tooltip_content.strip():
                try:
                    # Create a data validation with tooltip that shows on hover
                    from openpyxl.worksheet.datavalidation import DataValidation
                    
                    # Create a data validation that allows any input but shows tooltip
                    validation = DataValidation(
                        type="custom",
                        formula1="TRUE",  # Always true, so any input is valid
                        showErrorMessage=False,
                        showInputMessage=True,
                        promptTitle=tooltip_title,
                        prompt=tooltip_content
                    )
                    
                    # Add the validation to the cell
                    validation.add(cell)
                    worksheet = cell.parent
                    worksheet.add_data_validation(validation)
                    
                    print("ExcelWriter: Added hover tooltip to cell {}: '{}' - '{}'".format(
                        cell.coordinate, tooltip_title, tooltip_content))
                except Exception as e:
                    print("ExcelWriter: Error adding tooltip to cell {}: {}".format(cell.coordinate, str(e)))
                    # Fallback to comment if data validation fails
                    try:
                        from openpyxl.comments import Comment
                        cell.comment = Comment(tooltip_content, tooltip_title)
                        print("ExcelWriter: Added comment tooltip to cell {}: '{}' - '{}'".format(
                            cell.coordinate, tooltip_title, tooltip_content))
                    except Exception as e2:
                        print("ExcelWriter: Error adding comment tooltip to cell {}: {}".format(cell.coordinate, str(e2)))
    
    def _convert_border_style(self, style_index):
        """
        Converts a border style index from xlsxwriter format to the appropriate openpyxl border style.
        
        Border style indexes based on xlsxwriter documentation:
        0: None           1: Thin           2: Medium
        3: Dash           4: Dot            5: Thick
        6: Double         7: Hair           8: Medium Dash
        9: Dash Dot       10: Medium Dash Dot  11: Dash Dot Dot
        12: Medium Dash Dot Dot  13: SlantDash Dot
        
        Args:
            style_index: The border style index (int or string representation)
            
        Returns:
            str: The openpyxl border style
        """
        # Convert to int if it's a string
        if isinstance(style_index, str) and style_index.isdigit():
            style_index = int(style_index)
        
        # Handle direct style names
        if isinstance(style_index, str):
            return style_index.lower()
            
        # Map numeric indexes to openpyxl styles
        style_map = {
            0: None,            # None
            1: "thin",          # Continuous (thin)
            2: "medium",        # Continuous (medium)
            3: "dashed",        # Dash
            4: "dotted",        # Dot
            5: "thick",         # Continuous (thick)
            6: "double",        # Double
            7: "hair",          # Continuous (hair)
            8: "mediumDashed",  # Dash (medium)
            9: "dashDot",       # Dash Dot
            10: "mediumDashDot", # Dash Dot (medium)
            11: "dashDotDot",   # Dash Dot Dot
            12: "mediumDashDotDot", # Dash Dot Dot (medium)
            13: "slantDashDot"  # SlantDash Dot
        }
        
        return style_map.get(style_index)

    def _get_column_index(self, column_reference):
        """
        Converts a column reference to a 1-based column index.
        
        Args:
            column_reference: Either an integer (0-based index) or string (Excel column letter)
            
        Returns:
            int: 1-based column index for Excel, or None if invalid
        """
        try:
            if isinstance(column_reference, int):
                # Convert 0-based index to 1-based
                return column_reference + 1
            elif isinstance(column_reference, str):
                # Convert Excel column letter to 1-based index
                return openpyxl.utils.column_index_from_string(column_reference.upper())
            else:
                print("ExcelWriter: Invalid column reference type. Must be integer or string.")
                return None
        except ValueError as e:
            print("ExcelWriter: Invalid column reference - {}".format(str(e)))
            return None


    def _apply_freeze_panes(self, worksheet, freeze_row=None, freeze_column=None):
        """
        Applies freeze panes to the worksheet based on specified row and/or column.
        
        Args:
            worksheet (openpyxl.Worksheet): The worksheet to apply freeze panes to
            freeze_row (int, optional): 0-based row index to freeze (rows above this will be frozen)
            freeze_column (int or str, optional): Either:
                - 0-based column index to freeze (columns to the left will be frozen)
                - Excel-style column letter (e.g., 'A', 'B', 'AA')
        
        Returns:
            bool: True if freeze panes were applied successfully, False otherwise
        """
        if freeze_row is None and freeze_column is None:
            return True

        try:
            # Validate and convert row index (always 0-based integer)
            excel_row = None
            if freeze_row is not None:
                if not isinstance(freeze_row, int):
                    print("ExcelWriter: Invalid freeze_row value. Must be an integer.")
                    return False
                excel_row = freeze_row + 1  # Convert 0-based to 1-based for Excel
            
            # Validate and convert column reference (can be 0-based int or Excel letter)
            excel_col = None
            if freeze_column is not None:
                excel_col = self._get_column_index(freeze_column)  # This already returns 1-based index
                if excel_col is None:
                    return False
            
            if excel_row and excel_col:
                # Freeze both row and column
                freeze_cell = "{}{}".format(
                    openpyxl.utils.get_column_letter(excel_col),
                    excel_row
                )
                worksheet.freeze_panes = worksheet[freeze_cell]
                print("ExcelWriter: Froze panes at cell {} (row {} and column {})".format(
                    freeze_cell, 
                    freeze_row,
                    freeze_column if isinstance(freeze_column, str) else openpyxl.utils.get_column_letter(excel_col)
                ))
            elif excel_row:
                # Freeze row only
                freeze_cell = "A{}".format(excel_row)
                worksheet.freeze_panes = worksheet[freeze_cell]
                print("ExcelWriter: Froze row {} (cell {})".format(freeze_row, freeze_cell))
            elif excel_col:
                # Freeze column only
                freeze_cell = "{}1".format(openpyxl.utils.get_column_letter(excel_col))
                worksheet.freeze_panes = worksheet[freeze_cell]
                print("ExcelWriter: Froze column {} (cell {})".format(
                    freeze_column if isinstance(freeze_column, str) else openpyxl.utils.get_column_letter(excel_col),
                    freeze_cell
                ))
            
            return True
            
        except Exception as e:
            print("ExcelWriter: Error applying freeze panes - {}".format(str(e)))
            import traceback
            print(traceback.format_exc())
            return False


class ExcelUpdater(ExcelHandler):
    """
    A class to update existing Excel .xlsx files with new values.
    Supports updating cell values and appending new rows while preserving formatting.
    """
    def __init__(self) -> None:
        super().__init__()

    def run(self):
        """
        Updates an Excel file with new values from the update data and appends new rows.
        Preserves existing formatting while updating cell contents.
        """
        file_path = self.job_data.get("filepath", None)
        if not file_path:
            print("ExcelUpdater: No file path provided.")
            return
        
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)
            worksheet_name = self.job_data.get("worksheet", workbook.sheetnames[0])
            worksheet = workbook[worksheet_name]
            
            # Handle update data
            self._handle_update_data(worksheet)
            
            # Handle append data
            self._handle_append_data(worksheet)
            
            # Save the workbook
            workbook.save(file_path)
            print("ExcelUpdater: File successfully updated and saved to '{}'".format(file_path))
            self.out = {"status": "success", "filepath": file_path}
            
        except Exception as e:
            print("ExcelUpdater: Error updating Excel file - {}".format(str(e)))
            print (traceback.format_exc())
            return 
            self.out = {"status": "error", "message": str(e)}
            
        
        self.dump_data()

    def _handle_update_data(self, worksheet):
        """
        Updates existing cells with new values from update_data.
        
        Args:
            worksheet (openpyxl.Worksheet): The worksheet to update
        """
        update_data = self.job_data.get("data", {}).get("update_data", {})
        for value in update_data.values():
            cell_location_row = value.get("row", None)
            cell_location_col = value.get("column", None)
            cell_new_value = value.get("value", None)
            
            if cell_location_row is not None and cell_location_col is not None:
                # Get the cell (using 1-based indices directly)
                cell = worksheet.cell(row=cell_location_row, column=cell_location_col)
                cell_current_value = cell.value
                
                # Update if value has changed
                if cell_current_value != cell_new_value:
                    print("Updating cell {},{} from {} to {}".format(
                        cell_location_row, cell_location_col, 
                        cell_current_value, cell_new_value))
                    cell.value = cell_new_value

    def _handle_append_data(self, worksheet):
        """
        Appends new rows of data to the worksheet.
        
        Args:
            worksheet (openpyxl.Worksheet): The worksheet to append data to
        """
        append_data = self.job_data.get("data", {}).get("append_data", {})
        if not append_data:
            return
            

        for row_data in append_data.values():
            row = row_data.get("row", None)
            col = row_data.get("column", None)
            cell = worksheet.cell(row=row, column=col)
            cell.value = row_data.get("value", None)

if __name__ == "__main__":
    """
    Entry point for the script.
    """
    ExcelHandler().run()